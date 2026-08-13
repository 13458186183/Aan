"""
asset_tool.py - 村级资产清查表工具路由
=======================================
提供资产清查表标准模板下载、填报数据上传校验、按类别汇总统计导出等功能。

业务背景:
村级资产清查表用于农村集体资产清产核资，逐项登记资产类别/名称、
账面数与清查数，并计算盘盈盘亏。本工具帮助填报人快速校验数据、
按类别汇总生成统计报表。

依赖:
- openpyxl: 生成/解析 Excel（模板、校验、汇总）
"""
import io
import json
import logging
import re
import uuid
from datetime import datetime
from urllib.parse import quote
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, JSONResponse, Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from utils.file_check import validate_single_file

logger = logging.getLogger("FieldKit.AssetTool")

router = APIRouter(prefix="/api/asset", tags=["村级资产清查表"])

BASE_DIR = Path(__file__).resolve().parent.parent
TEMP_DIR = BASE_DIR / "temp_file"
PARSE_DIR = TEMP_DIR / "asset_parse"    # 解析后的填报数据 JSON
OUTPUT_DIR = TEMP_DIR / "asset_output"  # 汇总统计表

# ============================================================
# 资产类别配置（可扩展）
# ============================================================
ASSET_CATEGORIES = [
    "流动资产",
    "固定资产",
    "农业资产",
    "无形资产",
    "长期投资",
    "其他资产",
]

# 填报字段列（与模板表头一一对应）
COLUMNS = [
    ("seq", "序号"),
    ("category", "资产类别"),
    ("name", "资产名称"),
    ("unit", "计量单位"),
    ("book_qty", "账面数量"),
    ("book_amt", "账面金额(元)"),
    ("check_qty", "清查数量"),
    ("check_amt", "清查金额(元)"),
    ("surplus_qty", "盘盈数量"),
    ("surplus_amt", "盘盈金额(元)"),
    ("deficit_qty", "盘亏数量"),
    ("deficit_amt", "盘亏金额(元)"),
    ("remark", "备注"),
]

HEADER_ROW = 4  # 模板中表头所在行（1 标题、2 说明、3 空、4 表头）
TITLE_TEXT = "村级资产清查表"

# ============================================================
# 工具函数
# ============================================================

def _ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def _error(status_code: int, detail: str) -> JSONResponse:
    return JSONResponse(status_code=status_code, content={"ok": False, "error": detail})


def _ok(data: dict = None, message: str = "ok") -> dict:
    return {"ok": True, "message": message, "data": data or {}}


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d%H%M%S%f")


def _norm(s: str) -> str:
    """规范化文本：去空格、去全半角括号，转小写"""
    return re.sub(r"[\s　/\\|()（）\[\]【】]", "", str(s or "")).lower()


def _to_number(value) -> Optional[float]:
    """
    将单元格值转为数值，失败返回 None。
    兼容：数字、带千分位的字符串（1,234.5）、中文数字（一二三）不支持。
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "").replace("，", "")
    try:
        return float(text)
    except (ValueError, TypeError):
        return None


def _num_str(value) -> str:
    """数值格式化：整数不带小数点，其余保留 2 位"""
    if value is None:
        return ""
    f = float(value)
    if f == int(f):
        return str(int(f))
    return f"{f:.2f}"


def _row_label(record: Dict, r_idx: int) -> str:
    """行定位标签：优先用填报的「序号」，附加 Excel 实际行号便于核对"""
    seq = (record.get("raw") or {}).get("seq", "").strip()
    if seq:
        return f"序号{seq}（第{r_idx}行）"
    return f"第{r_idx}行"


# ============================================================
# 模板生成
# ============================================================

def build_template() -> Workbook:
    """
    生成标准资产清查表模板：
    - Sheet1「资产清查表」：标题 + 填表说明 + 表头 + 示例行 + 类别下拉验证
    - Sheet2「资产类别」：类别选项（供下拉引用）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "资产清查表"

    # ---- 样式常量 ----
    title_font = Font(bold=True, size=16, color="1E293B")
    note_font = Font(size=10, color="64748B")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="D89DB5")
    example_font = Font(size=10, color="94A3B8", italic=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- 1. 标题 ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    c = ws.cell(row=1, column=1, value=TITLE_TEXT)
    c.font = title_font
    c.alignment = center
    ws.row_dimensions[1].height = 30

    # ---- 2. 填表说明 ----
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    c = ws.cell(
        row=2, column=1,
        value="填写说明：1)「资产类别」请在右侧下拉中选择；2) 数量/金额仅填数字；"
              "3) 清查数 = 账面数 + 盘盈 - 盘亏（可选填盘盈/盘亏）；4) 备注可填写资产说明。示例行为浅灰色，可删除。",
    )
    c.font = note_font
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # ---- 3. 表头 ----
    for col_idx, (_, header) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[HEADER_ROW].height = 24

    # ---- 4. 示例行（2 行） ----
    examples = [
        ["1", "流动资产", "库存现金", "元", "5000", "5000", "5200", "5200", "200", "200", "", "", "现金盘点溢余"],
        ["2", "固定资产", "办公电脑", "台", "3", "12000", "3", "12000", "", "", "", "", "已折旧"],
    ]
    for r_offset, row in enumerate(examples, start=1):
        r = HEADER_ROW + r_offset
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = example_font
            cell.alignment = center if col_idx != 3 and col_idx != 13 else left
            cell.border = border

    # ---- 5. 类别下拉（数据验证，引用 Sheet2） ----
    dv = DataValidation(
        type="list",
        formula1="=资产类别!$A$1:$A$6",
        allow_blank=True,
        showDropDown=False,  # False 表示显示下拉箭头
        errorTitle="资产类别错误",
        error="请从下拉列表中选择资产类别",
        showErrorMessage=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"C{HEADER_ROW + 1}:C1000")

    # ---- 6. 列宽 ----
    widths = [6, 12, 18, 10, 12, 14, 12, 14, 10, 12, 10, 12, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet2：资产类别 ----
    ws2 = wb.create_sheet("资产类别")
    for i, cat in enumerate(ASSET_CATEGORIES, start=1):
        ws2.cell(row=i, column=1, value=cat)
    ws2.sheet_state = "visible"
    ws2.column_dimensions["A"].width = 12

    return wb


# ============================================================
# 填报数据解析与校验
# ============================================================

def _locate_header(rows: List[List[str]]) -> Optional[int]:
    """定位表头行：需同时包含「资产类别」与「资产名称」"""
    for i, row in enumerate(rows[:20]):
        cells = [_norm(c) for c in row]
        has_category = any("资产类别" == c or "类别" in c for c in cells)
        has_name = any("资产名称" == c or "名称" in c for c in cells)
        if has_category and has_name:
            return i
    return None


def _parse_workbook(content: bytes) -> List[List[str]]:
    """读取 xlsx 内容为二维字符串数组（跳过完全空行）"""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if any(cells):
            rows.append(cells)
    wb.close()
    return rows


def _validate_record(idx: int, record: Dict) -> List[str]:
    """
    校验单条资产记录，返回错误列表（空 = 通过）。
    record 字段: category,name,unit,book_qty,book_amt,check_qty,check_amt,
                 surplus_qty,surplus_amt,deficit_qty,deficit_amt,remark
    """
    errors: List[str] = []

    # 必填
    if not record.get("category"):
        errors.append("资产类别必填")
    elif record["category"] not in ASSET_CATEGORIES:
        errors.append(f"资产类别「{record['category']}」不在标准类别中")

    if not record.get("name"):
        errors.append("资产名称必填")

    # 数字字段校验
    for field, label in [
        ("book_qty", "账面数量"), ("book_amt", "账面金额"),
        ("check_qty", "清查数量"), ("check_amt", "清查金额"),
        ("surplus_qty", "盘盈数量"), ("surplus_amt", "盘盈金额"),
        ("deficit_qty", "盘亏数量"), ("deficit_amt", "盘亏金额"),
    ]:
        val = record.get(field)
        if val is not None and val < 0:
            errors.append(f"{label}不能为负数")

    # 盘盈/盘亏不能同时存在
    has_surplus = record.get("surplus_qty") not in (None, 0) or record.get("surplus_amt") not in (None, 0)
    has_deficit = record.get("deficit_qty") not in (None, 0) or record.get("deficit_amt") not in (None, 0)
    if has_surplus and has_deficit:
        errors.append("同一资产不能同时填写盘盈和盘亏")

    # 数量勾稽：清查数量 ≈ 账面数量 + 盘盈数量 - 盘亏数量（填写盘盈/盘亏数量时校验）
    if record.get("surplus_qty") not in (None, 0) or record.get("deficit_qty") not in (None, 0):
        if record.get("book_qty") is not None and record.get("check_qty") is not None:
            expect = record["book_qty"] + (record["surplus_qty"] or 0) - (record["deficit_qty"] or 0)
            if abs(expect - record["check_qty"]) > 1e-6:
                errors.append(
                    f"数量勾稽不符：清查数量应≈账面数量+盘盈-盘亏（{_num_str(expect)}）"
                )

    # 金额勾稽
    if record.get("surplus_amt") not in (None, 0) or record.get("deficit_amt") not in (None, 0):
        if record.get("book_amt") is not None and record.get("check_amt") is not None:
            expect = record["book_amt"] + (record["surplus_amt"] or 0) - (record["deficit_amt"] or 0)
            if abs(expect - record["check_amt"]) > 1e-6:
                errors.append(
                    f"金额勾稽不符：清查金额应≈账面金额+盘盈-盘亏（{_num_str(expect)}）"
                )

    # 序号行定位错误提示
    if errors:
        errors = [f"{_row_label(record, idx)}: " + "；".join(errors)]

    return errors


def _parse_asset_file(content: bytes, filename: str) -> Tuple[List[Dict], List[str]]:
    """
    解析资产清查表 xlsx，返回 (记录列表, 全局错误列表)。
    记录字段含数值已转 float 或 None；原始单元格保留 raw_ 前缀副本便于展示。
    """
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 格式的清查表")

    rows = _parse_workbook(content)
    if not rows:
        raise HTTPException(status_code=400, detail="文件内容为空")

    header_idx = _locate_header(rows)
    if header_idx is None:
        raise HTTPException(
            status_code=400,
            detail="未识别到表头：请使用模板填报，表头需包含「资产类别」「资产名称」",
        )

    headers = rows[header_idx]
    # 建立 标准字段名 -> 列索引 的映射（按模板列名匹配）
    col_map: Dict[str, Optional[int]] = {key: None for key, _ in COLUMNS}
    for key, label in COLUMNS:
        for i, h in enumerate(headers):
            if _norm(h) == _norm(label):
                col_map[key] = i
                break

    if col_map["category"] is None or col_map["name"] is None:
        raise HTTPException(
            status_code=400,
            detail="表头不完整：缺少「资产类别」或「资产名称」列",
        )

    records: List[Dict] = []
    all_errors: List[str] = []

    for r_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        def _cell(key: str) -> str:
            i = col_map[key]
            if i is None or i >= len(row):
                return ""
            return row[i]

        category = _cell("category")
        name = _cell("name")
        # 完全空行跳过
        if not category and not name and not any(_cell(k) for k, _ in COLUMNS):
            continue

        record: Dict = {
            "category": category,
            "name": name,
            "unit": _cell("unit"),
            "remark": _cell("remark"),
            # 数值字段
            "book_qty": _to_number(_cell("book_qty")),
            "book_amt": _to_number(_cell("book_amt")),
            "check_qty": _to_number(_cell("check_qty")),
            "check_amt": _to_number(_cell("check_amt")),
            "surplus_qty": _to_number(_cell("surplus_qty")),
            "surplus_amt": _to_number(_cell("surplus_amt")),
            "deficit_qty": _to_number(_cell("deficit_qty")),
            "deficit_amt": _to_number(_cell("deficit_amt")),
            # 原始字符串（用于展示）
            "raw": {
                "seq": _cell("seq"),
                "category": category,
                "name": name,
                "unit": _cell("unit"),
                "book_qty": _cell("book_qty"),
                "book_amt": _cell("book_amt"),
                "check_qty": _cell("check_qty"),
                "check_amt": _cell("check_amt"),
                "surplus_qty": _cell("surplus_qty"),
                "surplus_amt": _cell("surplus_amt"),
                "deficit_qty": _cell("deficit_qty"),
                "deficit_amt": _cell("deficit_amt"),
                "remark": _cell("remark"),
            },
        }

        # 数字格式错误检测（原始值非空但转不出数字）
        format_errors = []
        for key, label in [
            ("book_qty", "账面数量"), ("book_amt", "账面金额"),
            ("check_qty", "清查数量"), ("check_amt", "清查金额"),
            ("surplus_qty", "盘盈数量"), ("surplus_amt", "盘盈金额"),
            ("deficit_qty", "盘亏数量"), ("deficit_amt", "盘亏金额"),
        ]:
            raw = _cell(key)
            if raw and record[key] is None:
                format_errors.append(f"{label}「{raw}」不是有效数字")

        row_errors = _validate_record(r_idx, record)
        if format_errors:
            row_errors = [f"{_row_label(record, r_idx)}: " + "；".join(format_errors)] + row_errors
        record["errors"] = row_errors
        all_errors.extend(row_errors)
        records.append(record)

    if not records:
        raise HTTPException(status_code=400, detail="未解析到任何资产数据行")

    return records, all_errors


# ============================================================
# 汇总统计
# ============================================================

def _compute_summary(records: List[Dict]) -> Dict:
    """按资产类别汇总统计，返回 summary 数据"""
    cats: Dict[str, Dict] = {cat: {
        "count": 0,
        "book_amt": 0.0, "check_amt": 0.0,
        "surplus_amt": 0.0, "deficit_amt": 0.0,
        "has_check": 0,
    } for cat in ASSET_CATEGORIES}

    for r in records:
        cat = r.get("category")
        if cat not in cats:
            cat = "其他资产"
        cats[cat]["count"] += 1
        cats[cat]["book_amt"] += r.get("book_amt") or 0
        cats[cat]["check_amt"] += r.get("check_amt") or 0
        cats[cat]["surplus_amt"] += r.get("surplus_amt") or 0
        cats[cat]["deficit_amt"] += r.get("deficit_amt") or 0
        if r.get("check_amt") is not None:
            cats[cat]["has_check"] += 1

    total = {
        "count": sum(c["count"] for c in cats.values()),
        "book_amt": round(sum(c["book_amt"] for c in cats.values()), 2),
        "check_amt": round(sum(c["check_amt"] for c in cats.values()), 2),
        "surplus_amt": round(sum(c["surplus_amt"] for c in cats.values()), 2),
        "deficit_amt": round(sum(c["deficit_amt"] for c in cats.values()), 2),
    }

    # 差值：清查 - 账面（= 盘盈 - 盘亏）
    total["diff_amt"] = round(total["check_amt"] - total["book_amt"], 2)

    rows = []
    for cat in ASSET_CATEGORIES:
        c = cats[cat]
        if c["count"] == 0:
            continue
        rows.append({
            "category": cat,
            "count": c["count"],
            "book_amt": round(c["book_amt"], 2),
            "check_amt": round(c["check_amt"], 2),
            "surplus_amt": round(c["surplus_amt"], 2),
            "deficit_amt": round(c["deficit_amt"], 2),
            "diff_amt": round(c["check_amt"] - c["book_amt"], 2),
            "has_check": c["has_check"],
        })
    return {"rows": rows, "total": total}


def build_summary_workbook(records: List[Dict], source_name: str) -> Workbook:
    """生成汇总统计表：Sheet1 明细 + Sheet2 按类别汇总"""
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="5B8DEF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---- Sheet1: 清查明细 ----
    ws = wb.active
    ws.title = "清查明细"
    headers = ["序号", "资产类别", "资产名称", "计量单位",
               "账面数量", "账面金额(元)", "清查数量", "清查金额(元)",
               "盘盈数量", "盘盈金额(元)", "盘亏数量", "盘亏金额(元)", "备注"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for i, r in enumerate(records, start=1):
        ws.append([
            i,
            r.get("category", ""),
            r.get("name", ""),
            r.get("unit", ""),
            _num_str(r.get("book_qty")),
            _num_str(r.get("book_amt")),
            _num_str(r.get("check_qty")),
            _num_str(r.get("check_amt")),
            _num_str(r.get("surplus_qty")),
            _num_str(r.get("surplus_amt")),
            _num_str(r.get("deficit_qty")),
            _num_str(r.get("deficit_amt")),
            r.get("remark", ""),
        ])

    widths = [6, 12, 18, 10, 12, 14, 12, 14, 10, 12, 10, 12, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet2: 类别汇总 ----
    ws2 = wb.create_sheet("类别汇总")
    summary = _compute_summary(records)
    sum_headers = ["资产类别", "条目数", "账面金额(元)", "清查金额(元)",
                   "盘盈金额(元)", "盘亏金额(元)", "差额(清查-账面)"]
    ws2.append(sum_headers)
    for col_idx in range(1, len(sum_headers) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row in summary["rows"]:
        ws2.append([
            row["category"], row["count"],
            _num_str(row["book_amt"]), _num_str(row["check_amt"]),
            _num_str(row["surplus_amt"]), _num_str(row["deficit_amt"]),
            _num_str(row["diff_amt"]),
        ])

    t = summary["total"]
    ws2.append([
        "合计", t["count"],
        _num_str(t["book_amt"]), _num_str(t["check_amt"]),
        _num_str(t["surplus_amt"]), _num_str(t["deficit_amt"]),
        _num_str(t["diff_amt"]),
    ])
    last_row = ws2.max_row
    for col_idx in range(1, len(sum_headers) + 1):
        ws2.cell(row=last_row, column=col_idx).font = Font(bold=True)
        ws2.cell(row=last_row, column=col_idx).fill = PatternFill("solid", fgColor="FBF0F4")

    for i, w in enumerate([12, 8, 14, 14, 14, 14, 16], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    return wb


# ============================================================
# 接口 ①：下载标准模板
# ============================================================

@router.get("/template", summary="下载资产清查表标准模板（xlsx）")
async def download_template():
    """
    生成并返回标准模板 Excel：
    - Sheet1 资产清查表：标题/说明/表头/示例行 + 类别下拉
    - Sheet2 资产类别：类别选项
    """
    _ensure_dir(TEMP_DIR)
    filename = f"村级资产清查表模板_{datetime.now().strftime('%Y%m%d')}.xlsx"
    buf = io.BytesIO()
    wb = build_template()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"asset_template.xlsx\"; "
                f"filename*=UTF-8''{quote(filename)}"
            )
        },
    )


# ============================================================
# 接口 ②：上传并校验填报数据
# ============================================================

@router.post("/parse", summary="上传资产清查表并校验（xlsx/xls）")
async def parse_asset(
    file: UploadFile = File(..., description="资产清查表文件（.xlsx / .xls）"),
):
    """
    上传填报好的资产清查表，自动解析并校验：
    - 必填项（资产类别/资产名称）
    - 数字格式与取值范围
    - 类别合法性
    - 盘盈/盘亏逻辑与勾稽关系

    返回:
    - session_id: 本次填报会话 ID（生成汇总表时使用）
    - records: 每条资产数据（含 raw 原始值、errors 校验错误）
    - errors: 全部错误列表（空则校验通过）
    - stats: 基础统计（总条数、有效条数）
    """
    ok, msg = validate_single_file(
        filename=file.filename or "",
        file_size=file.size or 0,
        module="asset",
    )
    if not ok:
        return _error(400, msg)

    content = await file.read()
    records, all_errors = _parse_asset_file(content, file.filename)

    # 保存解析结果
    _ensure_dir(PARSE_DIR)
    session_id = _ts() + uuid.uuid4().hex[:6]
    payload = {
        "source_filename": file.filename,
        "records": records,
        "errors": all_errors,
    }
    (PARSE_DIR / f"{session_id}.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    valid_count = sum(1 for r in records if not r.get("errors"))
    stats = {
        "total": len(records),
        "valid": valid_count,
        "invalid": len(records) - valid_count,
    }

    # 视图数据（供前端展示）
    view = []
    for r in records:
        raw = r.get("raw", {})
        view.append({
            "seq": raw.get("seq", ""),
            "category": r.get("category", ""),
            "name": r.get("name", ""),
            "unit": raw.get("unit", ""),
            "book_qty": raw.get("book_qty", ""),
            "book_amt": raw.get("book_amt", ""),
            "check_qty": raw.get("check_qty", ""),
            "check_amt": raw.get("check_amt", ""),
            "surplus_qty": raw.get("surplus_qty", ""),
            "surplus_amt": raw.get("surplus_amt", ""),
            "deficit_qty": raw.get("deficit_qty", ""),
            "deficit_amt": raw.get("deficit_amt", ""),
            "remark": raw.get("remark", ""),
            "errors": r.get("errors", []),
        })

    if all_errors:
        message = f"校验完成：共 {len(records)} 条，{len(records) - valid_count} 条存在 {len(all_errors)} 处错误"
    else:
        message = f"校验通过：共 {len(records)} 条资产记录，全部正确"

    return _ok({
        "session_id": session_id,
        "source_filename": file.filename,
        "records": view,
        "errors": all_errors,
        "stats": stats,
        "categories": ASSET_CATEGORIES,
    }, message=message)


# ============================================================
# 接口 ③：生成汇总统计表
# ============================================================

@router.post("/export", summary="根据填报数据生成汇总统计表（xlsx）")
async def export_summary(
    session_id: str = Form(..., description="填报会话 ID"),
    title: str = Form(default="", description="村名/表名（可选）"),
):
    """
    读取会话中的填报数据，生成汇总统计 Excel：
    - Sheet1 清查明细：完整明细
    - Sheet2 类别汇总：按类别统计 + 合计行
    """
    parse_path = PARSE_DIR / f"{session_id}.json"
    if not parse_path.exists():
        return _error(400, "填报会话已失效，请重新上传清查表")

    payload = json.loads(parse_path.read_text(encoding="utf-8"))
    records = payload["records"]
    if not records:
        return _error(400, "无有效填报数据")

    # 用新的解析结果（防止 records 里含 errors 影响）——其实 records 已含 errors 字段，无碍
    safe_title = re.sub(r'[\\/:*?"<>|]', "_", title.strip())
    source_name = safe_title or payload.get("source_filename", "村级资产清查")

    _ensure_dir(OUTPUT_DIR)
    filename = f"资产清查汇总_{safe_title or '村级'}_{_ts()}.xlsx"
    out_path = OUTPUT_DIR / filename

    wb = build_summary_workbook(records, source_name)
    wb.save(out_path)

    summary = _compute_summary(records)
    t = summary["total"]
    valid_count = sum(1 for r in records if not r.get("errors"))
    size_kb = round(out_path.stat().st_size / 1024, 1)

    return _ok({
        "filename": filename,
        "download_url": f"/api/asset/file/{filename}",
        "summary": summary,
        "stats": {
            "total": len(records),
            "valid": valid_count,
            "invalid": len(records) - valid_count,
        },
        "total_amt": t,
        "size_kb": size_kb,
    }, message=f"汇总表已生成：{len(records)} 条资产，账面金额 {_num_str(t['book_amt'])} 元，清查金额 {_num_str(t['check_amt'])} 元")


# ============================================================
# 公共：文件下载
# ============================================================

@router.get("/file/{filename}", summary="下载生成的汇总统计表")
async def download_file(filename: str):
    """下载生成的汇总统计 Excel"""
    target = (OUTPUT_DIR / filename).resolve()
    if not target.is_relative_to(OUTPUT_DIR.resolve()) or not target.exists():
        raise HTTPException(status_code=404, detail="文件不存在或已过期")
    return FileResponse(
        path=str(target),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )
