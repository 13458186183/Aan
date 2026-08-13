"""
attendance_tool.py - 考勤点名工具路由
=====================================
提供花名册解析（xlsx/csv）、合照上传、点名结果导出（xlsx）等功能。

依赖:
- openpyxl: 解析/生成 Excel
- 标准库 csv: 解析 CSV（支持 utf-8 / gbk 编码）
"""
import csv
import io
import json
import logging
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.file_check import validate_single_file

logger = logging.getLogger("FieldKit.AttendanceTool")

router = APIRouter(prefix="/api/attendance", tags=["考勤点名"])

BASE_DIR = Path(__file__).resolve().parent.parent
TEMP_DIR = BASE_DIR / "temp_file"
ROSTER_DIR = TEMP_DIR / "attendance_roster"   # 解析后的花名册 JSON
PHOTO_DIR = TEMP_DIR / "attendance_photos"    # 上传的合照
OUTPUT_DIR = TEMP_DIR / "attendance_output"   # 导出的考勤表

# ============================================================
# 表头识别配置
# ============================================================
NAME_KEYS = ["姓名", "名字", "名称", "成员", "人员", "name", "student", "person"]
NO_KEYS = ["学号", "工号", "编号", "序号", "账号", "id", "no", "num", "number", "studentid", "code"]
DEPT_KEYS = ["班级", "部门", "单位", "学院", "专业", "院系", "组", "团队", "科室", "dept", "class", "group", "department", "team"]

# 状态 -> 中文名
STATUS_TEXT = {
    "present": "到场",
    "late": "迟到",
    "leave": "请假",
    "absent": "缺席",
    "": "未点名",
}

STATUS_FILL = {
    "present": "C6EFCE",  # 绿
    "late": "FFEB9C",     # 黄
    "leave": "BDD7EE",    # 蓝
    "absent": "FFC7CE",   # 红
    "": "F2F2F2",         # 灰
}


# ============================================================
# 工具函数
# ============================================================

def _ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def _error(status_code: int, detail: str) -> JSONResponse:
    return JSONResponse(status_code=status_code, content={"ok": False, "error": detail})


def _ok(data: dict = None, message: str = "ok") -> dict:
    return {"ok": True, "message": message, "data": data or {}}


def _ts() -> str:
    return datetime.now().strftime("%Y%m%d%H%M%S%f")


def _norm(s: str) -> str:
    """规范化表头文本：去空格、转小写"""
    return re.sub(r"[\s　/\\|()（）\[\]【】]", "", str(s or "")).lower()


def _match_headers(headers: List[str]) -> Dict[str, Optional[int]]:
    """
    根据表头行匹配各字段列索引。
    返回 {"name": idx, "no": idx, "dept": idx}，未匹配为 None
    """
    result: Dict[str, Optional[int]] = {"name": None, "no": None, "dept": None}
    for idx, h in enumerate(headers):
        norm_h = _norm(h)
        if result["name"] is None and any(k in norm_h for k in NAME_KEYS):
            result["name"] = idx
        if result["no"] is None and any(k in norm_h for k in NO_KEYS):
            result["no"] = idx
        if result["dept"] is None and any(k in norm_h for k in DEPT_KEYS):
            result["dept"] = idx
    return result


def _read_rows_xlsx(content: bytes) -> Tuple[List[List[str]], Optional[int]]:
    """
    读取 xlsx 内容，返回 (二维字符串数组, 可能的表头行索引)。
    自动跳过空行。
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows: List[List[str]] = []
    header_row: Optional[int] = None
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        cells = ["" if c is None else str(c).strip() for c in row]
        if not any(cells):
            continue
        if header_row is None and any(_norm(c) for c in cells):
            header_row = r_idx
        rows.append(cells)
    wb.close()
    return rows, header_row


def _read_rows_csv(content: bytes) -> Tuple[List[List[str]], Optional[int]]:
    """
    读取 CSV 内容（自动探测编码），返回 (二维字符串数组, 可能的表头行索引)。
    """
    raw = content
    decoded: Optional[str] = None
    for enc in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            decoded = raw.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if decoded is None:
        decoded = raw.decode("utf-8", errors="replace")

    reader = csv.reader(io.StringIO(decoded))
    rows: List[List[str]] = []
    header_row: Optional[int] = None
    for r_idx, row in enumerate(reader):
        cells = [str(c or "").strip() for c in row]
        if not any(cells):
            continue
        if header_row is None and any(_norm(c) for c in cells):
            header_row = r_idx
        rows.append(cells)
    return rows, header_row


def _parse_roster(content: bytes, filename: str) -> Tuple[List[Dict[str, str]], Dict[str, str]]:
    """
    解析花名册文件，返回 (人员列表, 列映射说明)。
    人员格式: {"name": str, "no": str, "dept": str}
    列映射说明: {"name": 原始表头, "no": 原始表头, "dept": 原始表头}
    """
    ext = Path(filename).suffix.lower()

    if ext == ".xlsx":
        rows, hint_row = _read_rows_xlsx(content)
    elif ext == ".csv":
        rows, hint_row = _read_rows_csv(content)
    else:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .csv 格式的花名册")

    if not rows:
        raise HTTPException(status_code=400, detail="花名册内容为空")

    # 定位表头行：优先用 hint_row，否则在前 8 行中找匹配
    search_zone = rows[:8]
    header_idx: Optional[int] = None
    matched_columns: Dict[str, Optional[int]] = {"name": None, "no": None, "dept": None}

    if hint_row is not None:
        match = _match_headers(rows[hint_row])
        if match["name"] is not None:
            header_idx = hint_row
            matched_columns = match

    if header_idx is None:
        for i, row in enumerate(search_zone):
            match = _match_headers(row)
            if match["name"] is not None:
                header_idx = i
                matched_columns = match
                break

    if header_idx is None:
        raise HTTPException(
            status_code=400,
            detail="未识别到表头：请在花名册中包含「姓名」列（可含「学号/工号」「班级/部门」列）",
        )

    headers = rows[header_idx]
    col_name, col_no, col_dept = matched_columns["name"], matched_columns["no"], matched_columns["dept"]

    col_map_text: Dict[str, str] = {
        "name": headers[col_name] if col_name is not None else "",
        "no": headers[col_no] if col_no is not None else "",
        "dept": headers[col_dept] if col_dept is not None else "",
    }

    # 读取数据行
    people: List[Dict[str, str]] = []
    seen = set()
    for row in rows[header_idx + 1:]:
        def _cell(idx: Optional[int]) -> str:
            if idx is None or idx >= len(row):
                return ""
            return row[idx]

        name = _cell(col_name).strip()
        if not name:
            continue

        no = _cell(col_no).strip()
        dept = _cell(col_dept).strip()

        # 去重（姓名+学号同时相同视为重复）
        key = f"{name}|{no}"
        if key in seen:
            continue
        seen.add(key)

        people.append({"name": name, "no": no, "dept": dept})

    if not people:
        raise HTTPException(status_code=400, detail="花名册中没有解析到有效的人员名单")

    return people, col_map_text


def _build_workbook(records: List[Dict], title: str) -> Workbook:
    """
    根据点名结果生成 Excel 考勤表。
    records 格式: {"name","no","dept","status","remark"}
    """
    wb = Workbook()

    # ---- Sheet1: 考勤明细 ----
    ws = wb.active
    ws.title = "考勤明细"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="5B8DEF")
    center = Alignment(horizontal="center", vertical="center")

    headers = ["序号", "姓名", "学号/工号", "班级/部门", "考勤状态", "备注"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    stats = {"present": 0, "late": 0, "leave": 0, "absent": 0, "": 0}
    for idx, r in enumerate(records, start=1):
        status = r.get("status", "")
        stats[status] = stats.get(status, 0) + 1
        row_data = [
            idx,
            r.get("name", ""),
            r.get("no", ""),
            r.get("dept", ""),
            STATUS_TEXT.get(status, "未点名"),
            r.get("remark", ""),
        ]
        ws.append(row_data)
        fill_color = STATUS_FILL.get(status, "F2F2F2")
        fill = PatternFill("solid", fgColor=fill_color)
        ws.cell(row=idx + 1, column=5).fill = fill

    # 列宽
    widths = [6, 14, 14, 16, 12, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Sheet2: 考勤汇总 ----
    ws2 = wb.create_sheet("考勤汇总")
    total = len(records)
    present_count = stats.get("present", 0)
    late_count = stats.get("late", 0)
    # 出勤率口径：到场 + 迟到 视为出勤
    present_rate = f"{(present_count + late_count) / total * 100:.1f}%" if total else "0%"

    summary_rows = [
        ("考勤名称", title),
        ("应到人数", total),
        ("实到人数（到场）", present_count),
        ("迟到", late_count),
        ("请假", stats.get("leave", 0)),
        ("缺席", stats.get("absent", 0)),
        ("未点名", stats.get("", 0)),
        ("出勤率（到场+迟到）", present_rate),
        ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for label, value in summary_rows:
        ws2.append([label, value])

    for col_idx in range(1, 3):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 18
    ws2.cell(row=1, column=1).font = Font(bold=True)
    ws2.cell(row=2, column=1).font = Font(bold=True)

    return wb


# ============================================================
# 接口 ①：上传并解析花名册
# ============================================================

@router.post("/roster", summary="上传并解析花名册（xlsx/csv）")
async def parse_roster(
    file: UploadFile = File(..., description="花名册文件（.xlsx / .csv）"),
):
    """
    上传花名册，自动识别「姓名 / 学号 / 班级」列并解析人员名单。

    返回:
    - session_id: 本次花名册会话 ID（导出考勤表时使用）
    - columns: 识别到的列映射
    - people: 人员列表 [{name, no, dept}]
    """
    ok, msg = validate_single_file(
        filename=file.filename or "",
        file_size=file.size or 0,
        module="attendance",
        allowed_extensions=["xlsx", "csv"],
    )
    if not ok:
        return _error(400, msg)

    content = await file.read()
    people, col_map = _parse_roster(content, file.filename)

    # 保存解析结果，供导出接口使用
    _ensure_dir(ROSTER_DIR)
    session_id = _ts() + uuid.uuid4().hex[:6]
    payload = {
        "source_filename": file.filename,
        "columns": col_map,
        "people": people,
    }
    (ROSTER_DIR / f"{session_id}.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # 生成列表视图：姓名 + 序号
    view = [
        {"name": p["name"], "no": p["no"], "dept": p["dept"]}
        for p in people
    ]

    return _ok({
        "session_id": session_id,
        "columns": col_map,
        "people": view,
        "total": len(people),
    }, message=f"解析成功，共识别 {len(people)} 人")


# ============================================================
# 接口 ②：上传合照（用于辅助人工核对）
# ============================================================

@router.post("/photos", summary="上传活动合照（辅助人工点名核对）")
async def upload_photos(
    files: List[UploadFile] = File(..., description="合照图片，可多张"),
):
    """
    上传活动合照，返回可直接展示的图片地址列表。
    照片仅用于人工核对，不参与自动识别。
    """
    if not files:
        return _error(400, "请至少上传一张照片")

    _ensure_dir(PHOTO_DIR)
    photo_list = []
    for f in files:
        ok, msg = validate_single_file(
            filename=f.filename or "",
            file_size=f.size or 0,
            module="attendance",
            allowed_extensions=["jpg", "jpeg", "png", "webp", "gif", "bmp"],
        )
        if not ok:
            return _error(400, msg)

        # 清理文件名，防路径穿越
        safe_name = re.sub(r'[\\/:*?"<>|]', "_", Path(f.filename or "photo").name)
        safe_name = re.sub(r"\s+", "_", safe_name).strip(" .")
        if not safe_name:
            safe_name = f"photo_{uuid.uuid4().hex[:6]}"

        stored_name = f"{_ts()}_{uuid.uuid4().hex[:6]}_{safe_name}"
        content = await f.read()
        (PHOTO_DIR / stored_name).write_bytes(content)

        photo_list.append({
            "name": safe_name,
            "url": f"/api/attendance/photo/{stored_name}",
        })

    return _ok({"photos": photo_list}, message=f"成功上传 {len(photo_list)} 张合照")


# ============================================================
# 接口 ③：导出考勤表
# ============================================================

@router.post("/export", summary="根据点名结果导出 Excel 考勤表")
async def export_attendance(
    session_id: str = Form(..., description="花名册会话 ID"),
    records: str = Form(..., description="点名结果 JSON 字符串：[{name,no,dept,status,remark}]"),
    title: str = Form(default="考勤记录", description="考勤名称（默认：考勤记录）"),
):
    """
    接收前端点名结果，生成 Excel 考勤表（明细 + 汇总两个 Sheet）。

    records JSON 示例:
    [
      {"name": "张三", "no": "2021001", "dept": "一班", "status": "present", "remark": ""},
      {"name": "李四", "no": "2021002", "dept": "一班", "status": "absent", "remark": "未到"}
    ]
    """
    # 解析 records
    try:
        recs = json.loads(records)
        if not isinstance(recs, list):
            raise ValueError("records 必须是数组")
    except (json.JSONDecodeError, ValueError) as e:
        return _error(400, f"点名结果格式错误: {e}")

    if not recs:
        return _error(400, "点名结果为空")

    # 校验 session
    roster_path = ROSTER_DIR / f"{session_id}.json"
    if not roster_path.exists():
        return _error(400, "花名册会话已失效，请重新上传花名册")

    safe_title = re.sub(r'[\\/:*?"<>|]', "_", title.strip()) or "考勤记录"

    _ensure_dir(OUTPUT_DIR)
    filename = f"考勤表_{safe_title}_{_ts()}.xlsx"
    out_path = OUTPUT_DIR / filename

    wb = _build_workbook(recs, safe_title)
    wb.save(out_path)

    size_kb = round(out_path.stat().st_size / 1024, 1)
    total = len(recs)
    present = sum(1 for r in recs if r.get("status") == "present")
    absent = sum(1 for r in recs if r.get("status") == "absent")

    return _ok({
        "filename": filename,
        "download_url": f"/api/attendance/file/{filename}",
        "total": total,
        "present": present,
        "absent": absent,
        "size_kb": size_kb,
    }, message=f"考勤表已生成：应到 {total} 人，实到 {present} 人，缺席 {absent} 人")


# ============================================================
# 公共：合照展示 / 文件下载
# ============================================================

@router.get("/photo/{filename}", summary="展示上传的合照")
async def show_photo(filename: str):
    """展示上传的合照（人工核对用）"""
    target = (PHOTO_DIR / filename).resolve()
    if not target.is_relative_to(PHOTO_DIR.resolve()) or not target.exists():
        raise HTTPException(status_code=404, detail="照片不存在或已过期")
    return FileResponse(path=str(target), media_type="image/jpeg")


@router.get("/file/{filename}", summary="下载生成的考勤表")
async def download_file(filename: str):
    """下载导出的考勤 Excel 文件"""
    target = (OUTPUT_DIR / filename).resolve()
    if not target.is_relative_to(OUTPUT_DIR.resolve()) or not target.exists():
        raise HTTPException(status_code=404, detail="文件不存在或已过期")
    return FileResponse(
        path=str(target),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )
